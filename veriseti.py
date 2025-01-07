import os
import cv2
import sqlite3
from tkinter import Tk, Toplevel, Label, Entry, Button, PhotoImage, messagebox, END
from openpyxl import load_workbook

# Ana pencere ayarları
ana_pencere = Tk()
ana_pencere.title("Yoklama Sistemi")
ana_pencere.configure(background='light blue')
ana_pencere.geometry("800x400")

# Görsel referansını kaybetmemek için global değişken
newUserPhoto = PhotoImage(file=os.path.join(os.getcwd(), "new_user.png")).subsample(2, 2)
yoklamaResim = newUserPhoto.subsample(2, 2)

# Veritabanı bağlantısı
baglanti = sqlite3.connect(os.path.join(os.getcwd(), "personeller.db"))
im = baglanti.cursor()
im.execute("""
    CREATE TABLE IF NOT EXISTS personeller (
        ADI TEXT, 
        SOYADI TEXT, 
        GMAIL TEXT, 
        PAROLA TEXT
    )
""")
baglanti.commit()

# Veritabanına yeni kayıt ekleme fonksiyonu
def OgrenciyiKaydet():
    if ad_field.get() == "" or soyad_field.get() == "" or ogrenci_no_field.get() == "":
        messagebox.showerror("Hata!", "Lütfen tüm alanları doldurun!")
        return

    try:
        adi = ad_field.get()
        soyadi = soyad_field.get()
        mail = f"{adi.lower()}.{soyadi.lower()}@gmail.com"
        parola = "default"

        kaydet_bilgiler(adi, soyadi, mail, parola)
        messagebox.showinfo("Başarılı", f"Kayıt başarıyla eklendi! \nAdı: {adi}, Soyadı: {soyadi}, Gmail: {mail}")

        ad_field.delete(0, END)
        soyad_field.delete(0, END)
        ogrenci_no_field.delete(0, END)
    except Exception as e:
        messagebox.showerror("Hata", f"Kayıt işlemi başarısız: {e}")

def kaydet_bilgiler(adi, soyadi, gmail, parola):
    try:
        im.execute("INSERT INTO personeller (ADI, SOYADI, GMAIL, PAROLA) VALUES (?, ?, ?, ?)",
                   (adi, soyadi, gmail, parola))
        baglanti.commit()
    except Exception as e:
        messagebox.showerror("Hata", f"Veritabanına kayıt eklenemedi: {e}")

# Yeni öğrenci kayıt penceresi
def YeniOgrenciKaydet():
    kayit_penceresi = Toplevel()
    kayit_penceresi.configure(background='light green')
    kayit_penceresi.title("Öğrenci Kayıt Formu")
    kayit_penceresi.geometry("500x300")

    Label(kayit_penceresi, text="Yeni Öğrenci Kayıt Et", bg="light green", font="Helvetica 18 bold").grid(row=0, column=1)
    Label(kayit_penceresi, text="Öğrenci No", bg="light green").grid(row=1, column=0)
    Label(kayit_penceresi, text="Ad", bg="light green").grid(row=2, column=0)
    Label(kayit_penceresi, text="Soyad", bg="light green").grid(row=3, column=0)

    global ogrenci_no_field, ad_field, soyad_field
    ogrenci_no_field = Entry(kayit_penceresi)
    ad_field = Entry(kayit_penceresi)
    soyad_field = Entry(kayit_penceresi)

    ogrenci_no_field.grid(row=1, column=1, ipadx="100")
    ad_field.grid(row=2, column=1, ipadx="100")
    soyad_field.grid(row=3, column=1, ipadx="100")

    wb = load_workbook(os.path.join(os.getcwd(), 'sinif_listesi.xlsx'))
    sheet = wb.active

    def OgrenciyiKaydet():
        if ad_field.get() == "" or soyad_field.get() == "" or ogrenci_no_field.get() == "":
            messagebox.showerror("Hata!", "Lütfen tüm alanları doldurun!")
            return

        try:
            current_row = sheet.max_row
            sheet.cell(row=current_row + 1, column=1).value = ogrenci_no_field.get()
            sheet.cell(row=current_row + 1, column=2).value = ad_field.get()
            sheet.cell(row=current_row + 1, column=3).value = soyad_field.get()
            wb.save(os.path.join(os.getcwd(), 'sinif_listesi.xlsx'))
            messagebox.showinfo("Başarılı", "Öğrenci başarıyla kayıt edildi!")

            ad_field.delete(0, END)
            soyad_field.delete(0, END)
            ogrenci_no_field.delete(0, END)
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt işlemi başarısız: {e}")

    def YuzAlma():
        try:
            cam = cv2.VideoCapture(0)
            if not cam.isOpened():
                raise Exception("Kamera açılamadı. Lütfen bağlantıyı kontrol edin.")

            detector = cv2.CascadeClassifier(os.path.join(os.getcwd(), 'haarcascade_frontalface_default.xml'))
            if detector.empty():
                raise FileNotFoundError("Cascade dosyası bulunamadı.")

            Id = ogrenci_no_field.get()
            sampleNum = 0

            while True:
                ret, img = cam.read()
                if not ret:
                    print("Kamera görüntüsü alınamadı.")
                    break

                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                faces = detector.detectMultiScale(gray, 1.3, 5)

                for (x, y, w, h) in faces:
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    sampleNum += 1
                    cv2.imwrite(f"datasets/{Id}.{sampleNum}.jpg", gray[y:y+h, x:x+w])
                    cv2.imshow('Yüz Tarama', img)

                if cv2.waitKey(1) & 0xFF == ord('q') or sampleNum >= 50:
                    break

            cam.release()
            cv2.destroyAllWindows()
            print(f"[INFO] {sampleNum} yüz görüntüsü kaydedildi.")

        except Exception as e:
            print(f"Hata: {e}")

    Button(kayit_penceresi, text="Kaydet", fg="White", bg="Red", command=OgrenciyiKaydet).grid(row=8, column=0, ipadx="50", ipady="20")
    Button(kayit_penceresi, text="Yüz Tara", command=YuzAlma).grid(row=8, column=1, ipadx="50", ipady="20")

    kayit_penceresi.mainloop()

Button(ana_pencere, text="Öğrenci kaydet", font="Helvetica 18 bold", image=newUserPhoto, compound='top',
       bg="light blue", fg="blue", command=YeniOgrenciKaydet).grid(column=0, row=5, padx=20, pady=20, ipadx=20, ipady=20, sticky="nsew")

Button(ana_pencere, text="Akademisyen kaydet", font="Helvetica 18 bold", image=yoklamaResim, compound='top',
       bg="light blue", fg="blue", command=lambda: kaydet_bilgiler("Ad", "Soyad", "Mail", "Parola")).grid(column=1, row=5, padx=20, pady=20, ipadx=20, ipady=20, sticky="nsew")

# Grid yapılandırması
ana_pencere.grid_columnconfigure(0, weight=1)
ana_pencere.grid_columnconfigure(1, weight=1)
ana_pencere.grid_rowconfigure(5, weight=1)

ana_pencere.mainloop()